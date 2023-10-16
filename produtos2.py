# preencher a planilha
####
# entrar no site: https://www.novaliderinformatica.com.br/computadores-padrao
from selenium import webdriver
from selenium.webdriver.common.by import By
import openpyxl

##################################################

#criando planilha
workbook = openpyxl.Workbook()

#criando páginas
workbook.create_sheet('Computador')
workbook.create_sheet('Update')

#colocando titulo
sheet_produtos = workbook['Computador'] 
sheet_produtos['A1'].value = 'Produto'
sheet_produtos['B1'].value = 'Preço'

#colocando titulo
sheet_produtos1 = workbook['Update'] 
sheet_produtos1['A1'].value = 'Produto'
sheet_produtos1['B1'].value = 'Preço'

##################################################

##########################
#acessar o site
driver = webdriver.Chrome()
driver.get("https://www.novaliderinformatica.com.br/computadores-padrao")

##### pegando dos computadores
# pegando o nome
titulos = driver.find_elements(By.XPATH, "//a[@class = 'nome-produto']")

# pegar o preço
precos = driver.find_elements(By.XPATH, "//strong[@class ='preco-promocional']")
###########################


###########################
driver1 = webdriver.Chrome()
driver1.get("https://www.novaliderinformatica.com.br/kit-upgrade")

###### pegando dos updates
#pegando nome
titulos1 = driver1.find_elements(By.XPATH, "//a[@class = 'nome-produto']")

#pegando preço
precos1 = driver1.find_elements(By.XPATH, "//strong[@class ='preco-promocional']")
###########################

#inserir na tabela 
for titulo, preco in zip(titulos, precos):
    sheet_produtos.append([titulo.text, preco.text])

for titulo1, preco1 in zip(titulos1, precos1):
    sheet_produtos1.append([titulo1.text, preco1.text])

workbook.save('produto.xlsx')
